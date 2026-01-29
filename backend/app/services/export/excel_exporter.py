"""
Advanced Excel exporter with multi-sheet support.

This module exports crawl data, SEO analysis, and reports to Excel
with multiple sheets, formatting, and charts.
"""

from typing import Dict, List, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils import get_column_letter
from io import BytesIO


class ExcelExporter:
    """
    Export SEO data to Excel with multiple sheets and formatting.
    
    Creates professional Excel reports with:
    - Multiple sheets for different data types
    - Formatting and styling
    - Charts and visualizations
    - Filters and freeze panes
    """
    
    def __init__(self):
        """Initialize Excel exporter."""
        self.workbook = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
    
    def export_crawl_report(
        self,
        pages: List[Dict],
        summary: Dict,
        issues: List[Dict],
        project_name: str,
    ) -> bytes:
        """
        Export complete crawl report to Excel.
        
        Args:
            pages: List of crawled pages.
            summary: Crawl summary statistics.
            issues: List of detected issues.
            project_name: Project name for the report.
        
        Returns:
            bytes: Excel file content.
        """
        # Sheet 1: Summary
        self._create_summary_sheet(summary, project_name)
        
        # Sheet 2: All Pages
        self._create_pages_sheet(pages)
        
        # Sheet 3: Issues
        self._create_issues_sheet(issues)
        
        # Sheet 4: Status Codes
        self._create_status_codes_sheet(pages)
        
        # Sheet 5: Redirects
        redirects = [p for p in pages if 300 <= p.get('status_code', 0) < 400]
        if redirects:
            self._create_redirects_sheet(redirects)
        
        # Sheet 6: Errors
        errors = [p for p in pages if p.get('status_code', 0) >= 400]
        if errors:
            self._create_errors_sheet(errors)
        
        # Sheet 7: Meta Data
        self._create_metadata_sheet(pages)
        
        # Sheet 8: Images
        self._create_images_sheet(pages)
        
        # Save to bytes
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _create_summary_sheet(self, summary: Dict, project_name: str):
        """Create summary sheet with key metrics."""
        ws = self.workbook.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = f"SEO Crawl Report - {project_name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Summary metrics
        row = 4
        metrics = [
            ("Total Pages Crawled", summary.get('total_pages', 0)),
            ("Successful Pages (2xx)", summary.get('success_count', 0)),
            ("Redirects (3xx)", summary.get('redirect_count', 0)),
            ("Client Errors (4xx)", summary.get('client_error_count', 0)),
            ("Server Errors (5xx)", summary.get('server_error_count', 0)),
            ("Average Response Time (ms)", summary.get('avg_response_time', 0)),
            ("Total Issues Found", summary.get('total_issues', 0)),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_pages_sheet(self, pages: List[Dict]):
        """Create sheet with all pages data."""
        ws = self.workbook.create_sheet("All Pages")
        
        # Headers
        headers = [
            "URL", "Status Code", "Title", "Meta Description",
            "H1", "Word Count", "Internal Links", "External Links",
            "Images", "Images without Alt", "Response Time (ms)", "Depth"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, page in enumerate(pages, 2):
            ws.cell(row=row_idx, column=1, value=page.get('url', ''))
            ws.cell(row=row_idx, column=2, value=page.get('status_code', 0))
            ws.cell(row=row_idx, column=3, value=page.get('title', ''))
            ws.cell(row=row_idx, column=4, value=page.get('meta_description', ''))
            
            # H1 (first one if multiple)
            h1_tags = page.get('h1_tags', [])
            ws.cell(row=row_idx, column=5, value=h1_tags[0] if h1_tags else '')
            
            ws.cell(row=row_idx, column=6, value=page.get('word_count', 0))
            ws.cell(row=row_idx, column=7, value=page.get('internal_links_count', 0))
            ws.cell(row=row_idx, column=8, value=page.get('external_links_count', 0))
            ws.cell(row=row_idx, column=9, value=page.get('images_count', 0))
            ws.cell(row=row_idx, column=10, value=page.get('images_without_alt', 0))
            ws.cell(row=row_idx, column=11, value=page.get('response_time_ms', 0))
            ws.cell(row=row_idx, column=12, value=page.get('depth', 0))
            
            # Color code status codes
            status_cell = ws.cell(row=row_idx, column=2)
            status_code = page.get('status_code', 0)
            if 200 <= status_code < 300:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif 300 <= status_code < 400:
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif status_code >= 400:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-filter
        ws.auto_filter.ref = f"A1:L{len(pages) + 1}"
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
    
    def _create_issues_sheet(self, issues: List[Dict]):
        """Create sheet with detected issues."""
        ws = self.workbook.create_sheet("Issues")
        
        # Headers
        headers = ["Severity", "Type", "URL", "Issue Description", "Recommendation"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        # Data rows
        for row_idx, issue in enumerate(issues, 2):
            ws.cell(row=row_idx, column=1, value=issue.get('severity', 'Warning'))
            ws.cell(row=row_idx, column=2, value=issue.get('type', ''))
            ws.cell(row=row_idx, column=3, value=issue.get('url', ''))
            ws.cell(row=row_idx, column=4, value=issue.get('description', ''))
            ws.cell(row=row_idx, column=5, value=issue.get('recommendation', ''))
            
            # Color code by severity
            severity_cell = ws.cell(row=row_idx, column=1)
            severity = issue.get('severity', '').lower()
            if severity == 'critical':
                severity_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif severity == 'error':
                severity_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40
        ws.freeze_panes = "A2"
    
    def _create_status_codes_sheet(self, pages: List[Dict]):
        """Create sheet with status code distribution."""
        ws = self.workbook.create_sheet("Status Codes")
        
        # Count status codes
        status_counts = {}
        for page in pages:
            code = page.get('status_code', 0)
            status_counts[code] = status_counts.get(code, 0) + 1
        
        # Headers
        ws['A1'] = "Status Code"
        ws['B1'] = "Count"
        ws['C1'] = "Percentage"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True)
        
        # Data
        total = len(pages)
        row = 2
        for code, count in sorted(status_counts.items()):
            ws[f'A{row}'] = code
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{(count / total * 100):.1f}%"
            row += 1
        
        # Add chart
        chart = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=row-1)
        data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Status Code Distribution"
        ws.add_chart(chart, "E2")
    
    def _create_redirects_sheet(self, redirects: List[Dict]):
        """Create sheet with redirect analysis."""
        ws = self.workbook.create_sheet("Redirects")
        
        headers = ["URL", "Status Code", "Redirect Type", "Response Time (ms)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row_idx, page in enumerate(redirects, 2):
            ws.cell(row=row_idx, column=1, value=page.get('url', ''))
            ws.cell(row=row_idx, column=2, value=page.get('status_code', 0))
            
            code = page.get('status_code', 0)
            if code == 301:
                redirect_type = "Permanent (301)"
            elif code == 302:
                redirect_type = "Temporary (302)"
            elif code == 307:
                redirect_type = "Temporary (307)"
            elif code == 308:
                redirect_type = "Permanent (308)"
            else:
                redirect_type = f"Other ({code})"
            
            ws.cell(row=row_idx, column=3, value=redirect_type)
            ws.cell(row=row_idx, column=4, value=page.get('response_time_ms', 0))
        
        ws.column_dimensions['A'].width = 60
    
    def _create_errors_sheet(self, errors: List[Dict]):
        """Create sheet with error pages."""
        ws = self.workbook.create_sheet("Errors")
        
        headers = ["URL", "Status Code", "Error Type", "Depth"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        for row_idx, page in enumerate(errors, 2):
            ws.cell(row=row_idx, column=1, value=page.get('url', ''))
            ws.cell(row=row_idx, column=2, value=page.get('status_code', 0))
            
            code = page.get('status_code', 0)
            if code == 404:
                error_type = "Not Found (404)"
            elif code == 403:
                error_type = "Forbidden (403)"
            elif code >= 500:
                error_type = f"Server Error ({code})"
            else:
                error_type = f"Client Error ({code})"
            
            ws.cell(row=row_idx, column=3, value=error_type)
            ws.cell(row=row_idx, column=4, value=page.get('depth', 0))
        
        ws.column_dimensions['A'].width = 60
    
    def _create_metadata_sheet(self, pages: List[Dict]):
        """Create sheet with meta tag analysis."""
        ws = self.workbook.create_sheet("Meta Data")
        
        headers = [
            "URL", "Title", "Title Length", "Meta Description",
            "Meta Description Length", "Canonical URL", "OG Tags"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row_idx, page in enumerate(pages, 2):
            title = page.get('title', '')
            desc = page.get('meta_description', '')
            
            ws.cell(row=row_idx, column=1, value=page.get('url', ''))
            ws.cell(row=row_idx, column=2, value=title)
            ws.cell(row=row_idx, column=3, value=len(title))
            ws.cell(row=row_idx, column=4, value=desc)
            ws.cell(row=row_idx, column=5, value=len(desc))
            ws.cell(row=row_idx, column=6, value=page.get('canonical_url', ''))
            
            og_tags = page.get('og_tags', {})
            ws.cell(row=row_idx, column=7, value=len(og_tags) if og_tags else 0)
            
            # Highlight issues
            if not title:
                ws.cell(row=row_idx, column=2).fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
            if not desc:
                ws.cell(row=row_idx, column=4).fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['D'].width = 50
    
    def _create_images_sheet(self, pages: List[Dict]):
        """Create sheet with image analysis."""
        ws = self.workbook.create_sheet("Images")
        
        headers = [
            "URL", "Total Images", "Images without Alt",
            "Alt Coverage %", "Issue Severity"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row_idx, page in enumerate(pages, 2):
            total_images = page.get('images_count', 0)
            images_no_alt = page.get('images_without_alt', 0)
            
            if total_images > 0:
                coverage = ((total_images - images_no_alt) / total_images) * 100
            else:
                coverage = 100
            
            ws.cell(row=row_idx, column=1, value=page.get('url', ''))
            ws.cell(row=row_idx, column=2, value=total_images)
            ws.cell(row=row_idx, column=3, value=images_no_alt)
            ws.cell(row=row_idx, column=4, value=f"{coverage:.1f}%")
            
            # Severity
            if coverage < 50:
                severity = "Critical"
                ws.cell(row=row_idx, column=5).fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
            elif coverage < 80:
                severity = "Warning"
                ws.cell(row=row_idx, column=5).fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )
            else:
                severity = "Good"
                ws.cell(row=row_idx, column=5).fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
            
            ws.cell(row=row_idx, column=5, value=severity)
        
        ws.column_dimensions['A'].width = 60


def export_to_excel(
    pages: List[Dict],
    summary: Dict,
    issues: List[Dict],
    project_name: str
) -> bytes:
    """
    Convenience function to export crawl data to Excel.
    
    Args:
        pages: List of crawled pages.
        summary: Summary statistics.
        issues: List of issues.
        project_name: Project name.
    
    Returns:
        bytes: Excel file content.
    """
    exporter = ExcelExporter()
    return exporter.export_crawl_report(pages, summary, issues, project_name)
